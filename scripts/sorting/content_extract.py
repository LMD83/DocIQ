#!/usr/bin/env python3
"""
content_extract.py v4 — Tier A (deeper regex) + free Tesseract OCR fallback.

Extracts 25 fields per document from PDF / DOCX / XLSX, using text extraction
first and falling back to OCR (free, offline, Tesseract) when the text layer
is empty. Writes to ContentExtraction.csv with stable columns so
sort_folders.py v3 can score against the Tier A fields directly.

Tier A fields (v3 additions on top of the original 14):
  monetary, dac_num, bcar_num, planning_num, abp_num, rev_code, doc_status,
  reg_refs, author_firm, certifier, floor_area_m2, bed_count

OCR fallback (v4 addition):
  When pypdf returns no text for a PDF page, we render the page to an image
  via pdf2image and run pytesseract on it. Controlled by --ocr-pass and
  --max-ocr-pages to keep runtime sane.

Dependencies:
  Free:  pypdf, python-docx, openpyxl, pytesseract, pdf2image, Pillow
  System:
    - Tesseract OCR binary installed (Windows installer from UB Mannheim build)
    - Poppler utilities for pdf2image (on Windows: add poppler bin/ to PATH)

Usage:
  py content_extract.py                 # Tier A regex only (no OCR)
  py content_extract.py --ocr-pass      # run OCR on pages with no extractable text
  py content_extract.py --ocr-retry     # retry rows currently marked "no_text"/"skip_stuck"
  py content_extract.py --max-ocr-pages 2  # limit OCR work per file

Paths (edit if your layout differs):
"""
import argparse
import csv
import json
import os
import re
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

# ---------------------------------------------------------------------------
# Paths — override via env vars if your layout differs
# ---------------------------------------------------------------------------
DNREF_ROOT = Path(os.environ.get("DNREF_ROOT", r"C:\Users\gavin\Downloads\fwdeets\DNREF"))
CODES_DIR = Path(os.environ.get("GOVIQ_CODES", r"C:\Users\gavin\Downloads\GovIQ-Codes"))
INVENTORY_CSV = CODES_DIR / "FileInventory.csv"
RESULTS_CSV = CODES_DIR / "ContentExtraction.csv"
BUILDING_REG = CODES_DIR / "BuildingRegister.csv"

PER_FILE_SEC = 20
WORKERS = 4
PROGRESS_EVERY = 10
OCR_DEFAULT_MAX_PAGES = 2  # conservative default when OCR is enabled

# ---------------------------------------------------------------------------
# Worker script — runs in an isolated subprocess per file so one bad PDF
# can't crash the main loop
# ---------------------------------------------------------------------------
WORKER_SCRIPT = '''
import sys, json, os
job = json.loads(sys.stdin.read())
path = job["path"]; ext = job["ext"]
ocr_enabled = bool(job.get("ocr", False))
ocr_max_pages = int(job.get("ocr_max_pages", 2))

text = ""
used_ocr = False
try:
    if ext == "pdf":
        import pypdf
        r = pypdf.PdfReader(path, strict=False)
        for p in r.pages[:3]:
            try:
                text += (p.extract_text() or "")[:4000] + "\\n"
            except Exception:
                pass
        # OCR fallback — only if text layer was empty and OCR is enabled
        if ocr_enabled and not text.strip():
            try:
                from pdf2image import convert_from_path
                import pytesseract
                images = convert_from_path(path, dpi=200, first_page=1, last_page=ocr_max_pages)
                for img in images:
                    text += pytesseract.image_to_string(img)[:4000] + "\\n"
                if text.strip():
                    used_ocr = True
            except Exception as e:
                # OCR failure is non-fatal — we just stay with empty text
                pass
    elif ext == "docx":
        import docx
        d = docx.Document(path)
        text = "\\n".join(p.text for p in d.paragraphs[:200])
    elif ext == "xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames[:3]:
            for i, row in enumerate(wb[sheet].iter_rows(values_only=True)):
                if i > 60: break
                text += " ".join(str(v) for v in row if v is not None) + "\\n"
except Exception as e:
    print(json.dumps({"error": str(e)[:200], "text": "", "used_ocr": False}))
    sys.exit(0)
print(json.dumps({"text": text[:12000], "used_ocr": used_ocr}))
'''

# ---------------------------------------------------------------------------
# Patterns — Tier A
# ---------------------------------------------------------------------------
DN_PAT = re.compile(r"(DN|MH)[\s\-]?(\d{3,6})([A-Z])?", re.I)
DWG_PAT = re.compile(r"\b([A-Z]{1,4}[\-\/]\d{2,5}[\-\/]?[A-Z]?\d{0,3})\b")
DATE_PAT = re.compile(r"\b(?:(?:19|20)\d{2}|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{4})\b")
FSC_PAT = re.compile(r"\bF(?:SC|A)[\s\/\-]?\d+[\/\-]?\d+", re.I)
EIRCODE = re.compile(r"\b[A-Z]\d{2}\s?[A-Z0-9]{4}\b")

MONEY_PAT = re.compile(r"€\s*[\d,]+(?:\.\d{2})?(?:[\s\-]\s*€?\s*[\d,]+)?")
DAC_NUM = re.compile(r"\bDAC[\s\-\/]*\d+", re.I)
BCAR_NUM = re.compile(r"\bBCAR[\s\-\/]*[A-Z0-9\-\/]+", re.I)
PLANNING_NUM = re.compile(r"\b\d{3,5}\s*\/\s*\d{2}\b")
ABP_NUM = re.compile(r"\bABP[\s\-\/]?\d+", re.I)
REV_PAT = re.compile(r"\b(?:Rev(?:ision)?\.?|Issue)\s+([A-Z][\-\d]*|\d{1,2})", re.I)
STATUS_PAT = re.compile(
    r"\b(issued|approved|rejected|superseded|active|outstanding|completed|draft|"
    r"for\s+(?:review|approval|construction|tender)|granted|non[\s\-]conformance)\b",
    re.I,
)
REG_PAT = re.compile(
    r"\b(?:Part\s+[A-M]|Section\s+\d+|HIQA\s+Reg(?:ulation)?\s+\d+|BCAR|Building\s+Regulations|Article\s+\d+)\b",
    re.I,
)
AUTHOR_PAT = re.compile(
    r"(?:prepared\s+by|authored\s+by|issued\s+by|by\s*:|surveyed\s+by|inspected\s+by)[\s:]+"
    r"([A-Z][A-Za-z\s&\,\.\-']{2,60}?)"
    r"(?:\s+(?:Ltd|Limited|Partnership|LLP|Consultants|Engineers|Architects|Associates)|\n|$|\.)",
    re.I,
)
CERTIFIER_PAT = re.compile(
    r"(?:certified\s+by|assigned\s+certifier|certifier)[\s:]+([A-Z][A-Za-z\s\.\-']{2,40})",
    re.I,
)
FLOOR_AREA = re.compile(r"(\d+(?:,\d+)?(?:\.\d+)?)\s*(?:m²|m2|sq\.?\s*m|square\s+met)", re.I)
BEDS_PAT = re.compile(r"\b(\d+)\s*(?:bed(?:room|s)?|beds)\b", re.I)


# ---------------------------------------------------------------------------
# Extraction result stamper
# ---------------------------------------------------------------------------

EMPTY_TIER_A = ("",) * 17  # placeholder tuple for error paths


def extract_signals(text: str, known_dns: set, bnames: list):
    """Run all regex + building-name fuzz against the extracted text."""
    # Original signals
    dns = list(set(
        f"{m.group(1).upper()}{m.group(2).zfill(4)}{(m.group(3) or '').upper()}"
        for m in DN_PAT.finditer(text)
    ))[:5]

    matched = ""
    tl = text.lower()
    for orig, lc in bnames:
        if lc in tl:
            matched = orig
            break

    dwg = list(dict.fromkeys(DWG_PAT.findall(text)))[:3]
    dates = list(dict.fromkeys(DATE_PAT.findall(text)))[:3]
    fsc = list(dict.fromkeys(FSC_PAT.findall(text)))[:2]
    eir = list(dict.fromkeys(EIRCODE.findall(text)))[:2]

    # Tier A
    money = list(dict.fromkeys(MONEY_PAT.findall(text)))[:3]
    dac = list(dict.fromkeys(DAC_NUM.findall(text)))[:2]
    bcar = list(dict.fromkeys(BCAR_NUM.findall(text)))[:2]
    planning = list(dict.fromkeys(PLANNING_NUM.findall(text)))[:3]
    abp = list(dict.fromkeys(ABP_NUM.findall(text)))[:2]
    rev = list(dict.fromkeys(REV_PAT.findall(text)))[:2]
    status = list(dict.fromkeys([s.lower() for s in STATUS_PAT.findall(text)]))[:3]
    regs = list(dict.fromkeys(REG_PAT.findall(text)))[:3]
    author_m = AUTHOR_PAT.search(text)
    author = author_m.group(1).strip() if author_m else ""
    certifier_m = CERTIFIER_PAT.search(text)
    certifier = certifier_m.group(1).strip() if certifier_m else ""
    area_m = FLOOR_AREA.search(text)
    floor_area = area_m.group(1) if area_m else ""
    beds_m = BEDS_PAT.search(text)
    beds = beds_m.group(1) if beds_m else ""

    return {
        "dns": dns, "matched": matched, "dwg": dwg, "dates": dates, "fsc": fsc, "eir": eir,
        "money": money, "dac": dac, "bcar": bcar, "planning": planning, "abp": abp,
        "rev": rev, "status": status, "regs": regs,
        "author": author, "certifier": certifier, "floor_area": floor_area, "beds": beds,
    }


def compute_boost(signals: dict, site: str, known_dns: set) -> float:
    boost = 0.0
    if signals["dns"]: boost += 0.20
    if signals["matched"]: boost += 0.15
    if signals["dwg"]: boost += 0.05
    if signals["dates"]: boost += 0.05
    if signals["fsc"] or signals["dac"] or signals["bcar"] or signals["planning"] or signals["abp"]:
        boost += 0.10
    if signals["author"] or signals["certifier"]:
        boost += 0.05
    if signals["regs"]:
        boost += 0.05
    if site in known_dns:
        boost += 0.05
    if signals["money"]:
        boost += 0.02
    if signals["rev"]:
        boost += 0.03
    return round(min(0.80, boost), 3)


def run_one(job, known, bnames, ocr_enabled: bool, ocr_max_pages: int):
    site, folder, rel, fname, ext, size = job
    path = DNREF_ROOT / folder / rel
    size_kb = round(size / 1024, 1)

    if not path.exists():
        return (site, rel, fname, ext, size_kb) + EMPTY_TIER_A + (0.0, 0, "no", "missing")

    t0 = time.time()
    try:
        proc = subprocess.run(
            [sys.executable, "-c", WORKER_SCRIPT],
            input=json.dumps({
                "path": str(path),
                "ext": ext,
                "ocr": ocr_enabled,
                "ocr_max_pages": ocr_max_pages,
            }),
            capture_output=True, text=True, timeout=PER_FILE_SEC,
            creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0,
        )
        ms = int((time.time() - t0) * 1000)
        if proc.returncode != 0:
            return (site, rel, fname, ext, size_kb) + EMPTY_TIER_A + (0.0, ms, "no", "subproc_err")
        try:
            result = json.loads(proc.stdout)
        except Exception:
            return (site, rel, fname, ext, size_kb) + EMPTY_TIER_A + (0.0, ms, "no", "json_err")
        text = result.get("text", "")
        used_ocr = bool(result.get("used_ocr"))
    except subprocess.TimeoutExpired:
        return (site, rel, fname, ext, size_kb) + EMPTY_TIER_A + (0.0, PER_FILE_SEC * 1000, "no", "timeout")
    except Exception:
        return (site, rel, fname, ext, size_kb) + EMPTY_TIER_A + (0.0, int((time.time() - t0) * 1000), "no", "err")

    if not text:
        return (site, rel, fname, ext, size_kb) + EMPTY_TIER_A + (0.0, ms, "no", "no_text")

    ms = int((time.time() - t0) * 1000)
    s = extract_signals(text, known, bnames)
    boost = compute_boost(s, site, known)

    return (
        site, rel, fname, ext, size_kb,
        "; ".join(s["dns"]), s["matched"], "; ".join(s["dwg"]), "; ".join(s["dates"]),
        "; ".join(s["fsc"]), "; ".join(s["eir"]),
        "; ".join(s["money"]), "; ".join(s["dac"]), "; ".join(s["bcar"]),
        "; ".join(s["planning"]), "; ".join(s["abp"]), "; ".join(s["rev"]),
        "; ".join(s["status"]), "; ".join(s["regs"]),
        s["author"][:80], s["certifier"][:80], s["floor_area"], s["beds"],
        boost, ms, "yes" if used_ocr else "no", "ok",
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

COLS = [
    "site_code", "relative_path", "file_name", "extension", "size_kb",
    "dn_codes_in_content", "matched_building_name", "dwg_numbers", "dates", "fsc_refs", "eircodes",
    "monetary", "dac_num", "bcar_num", "planning_num", "abp_num",
    "rev_code", "doc_status", "reg_refs",
    "author_firm", "certifier", "floor_area_m2", "bed_count",
    "content_score_boost", "extraction_ms", "used_ocr", "status",
]

RETRY_STATUSES = {"no_text", "skip_stuck", "timeout", "subproc_err", "json_err", "err"}


def main():
    ap = argparse.ArgumentParser(description="Tier A content extraction with optional OCR fallback.")
    ap.add_argument("--ocr-pass", action="store_true", help="Enable Tesseract OCR fallback when text layer is empty.")
    ap.add_argument("--ocr-retry", action="store_true", help="Re-run OCR on rows currently marked with failure/empty statuses.")
    ap.add_argument("--max-ocr-pages", type=int, default=OCR_DEFAULT_MAX_PAGES, help="Max pages to OCR per PDF (default: 2).")
    args = ap.parse_args()

    ocr_enabled = args.ocr_pass or args.ocr_retry

    print("=" * 60)
    print(f"GovIQ Content Extraction v4 — Tier A{' + OCR' if ocr_enabled else ''}")
    print("=" * 60)
    start = time.time()

    known = set()
    bn = []
    if BUILDING_REG.exists():
        with open(BUILDING_REG, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if row.get("canonical_code"):
                    known.add(row["canonical_code"])
                nm = row.get("name", "")
                if nm and len(nm) > 6:
                    bn.append((nm, nm.lower()))

    processed = {}  # key -> status
    if RESULTS_CSV.exists():
        with open(RESULTS_CSV, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                key = f"{row['site_code']}::{row['relative_path']}"
                processed[key] = (row.get("status", "") or "").lower()

    jobs = []
    retryable_keys = set()
    if args.ocr_retry:
        retryable_keys = {k for k, s in processed.items() if s in RETRY_STATUSES}
        print(f"OCR retry mode: {len(retryable_keys)} rows eligible.")

    with open(INVENTORY_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row["extension"] not in ("pdf", "docx", "xlsx"):
                continue
            if row.get("dn_in_filename", ""):
                continue
            key = f"{row['site_code']}::{row['relative_path']}"
            if args.ocr_retry:
                if key not in retryable_keys:
                    continue
            else:
                if key in processed:
                    continue
            jobs.append((row["site_code"], row["folder_name"], row["relative_path"],
                         row["file_name"], row["extension"], int(row["size_bytes"])))

    jobs.sort(key=lambda j: j[5])
    print(f"Files to process: {len(jobs)}")
    print(f"Already in results: {len(processed)}")
    if not jobs:
        print("Nothing to do.")
        return

    mode = "a" if RESULTS_CSV.exists() and RESULTS_CSV.stat().st_size > 0 else "w"
    fh = open(RESULTS_CSV, mode, newline="", encoding="utf-8")
    w = csv.writer(fh)
    if mode == "w":
        w.writerow(COLS)

    cnt = 0
    timeouts = 0
    ocr_hits = 0
    print("Starting extraction ...")
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futures = {
            ex.submit(run_one, job, known, bn, ocr_enabled, args.max_ocr_pages): job
            for job in jobs
        }
        for fut in as_completed(futures):
            try:
                result = fut.result()
            except Exception:
                continue
            if result is None:
                continue
            w.writerow(result)
            fh.flush()
            cnt += 1
            if result[-1] == "timeout":
                timeouts += 1
            if result[-2] == "yes":
                ocr_hits += 1
            if cnt % PROGRESS_EVERY == 0:
                elapsed = time.time() - start
                rate = cnt / elapsed if elapsed > 0 else 0
                eta = (len(jobs) - cnt) / rate if rate > 0 else 0
                print(
                    f"  {cnt}/{len(jobs)} ({100 * cnt / len(jobs):.1f}%)  "
                    f"rate: {rate:.1f}/s  ETA: {int(eta // 60)}m{int(eta % 60)}s  "
                    f"timeouts: {timeouts}  ocr: {ocr_hits}"
                )
    fh.close()

    elapsed = time.time() - start
    print(
        f"\nDone. {cnt} files processed in {int(elapsed // 60)}m{int(elapsed % 60)}s "
        f"({timeouts} timeouts, {ocr_hits} OCR hits)"
    )


if __name__ == "__main__":
    main()
