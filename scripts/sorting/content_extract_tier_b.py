#!/usr/bin/env python3
"""
content_extract_tier_b.py — PAID add-on: Claude structured extraction.

For every document with an OK text-extraction row in ContentExtraction.csv,
send the extracted text to Claude with a structured-output tool schema and
record 23 semantic fields to ContentExtractionTierB.csv:

  document_type, scope_summary, primary_party, author_firm, certifier,
  approved_by, compliance_assertions, issues_identified, date_issued,
  date_expiry, date_superseded, monetary_amounts, referenced_documents,
  building_address, floor_area_m2, storeys, bed_count, rooms_named,
  regulatory_references, site_names, project_references,
  confidence, notes

This is OPTIONAL and costs money. At Claude Haiku rates, 13,000 docs x
~€0.02/doc = ~€260 one-off. Resumable. Safe to cancel and resume — already-
processed rows are skipped on re-run.

Usage:
  set ANTHROPIC_API_KEY=sk-ant-...
  py content_extract_tier_b.py              # full run
  py content_extract_tier_b.py --limit 100  # test on 100 docs first
  py content_extract_tier_b.py --dry-run    # cost estimate only

Zero-retention: configure zero-retention on the API key at
console.anthropic.com before processing customer data.
"""
import argparse
import csv
import json
import os
import sys
import time
from pathlib import Path

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
DNREF_ROOT = Path(os.environ.get("DNREF_ROOT", r"C:\Users\gavin\Downloads\fwdeets\DNREF"))
CODES_DIR = Path(os.environ.get("GOVIQ_CODES", r"C:\Users\gavin\Downloads\GovIQ-Codes"))
INVENTORY_CSV = CODES_DIR / "FileInventory.csv"
EXTRACT_CSV = CODES_DIR / "ContentExtraction.csv"
RESULTS_CSV = CODES_DIR / "ContentExtractionTierB.csv"

COST_PER_DOC_EUR = 0.020  # Haiku ballpark; refine from real usage

SYSTEM_PROMPT = """You are an HSE estates document analyst. You are given the first pages of
a construction or compliance document (extracted text, possibly noisy OCR).

Return a structured tool-use response with concise, factual fields. If a
field is not present in the text, return the empty string (or empty array).
Do not guess. Prefer "unclear" over fabrication."""

TOOL = {
    "name": "record_document",
    "description": "Structured semantic record for an HSE estates document.",
    "input_schema": {
        "type": "object",
        "properties": {
            "document_type": {
                "type": "string",
                "description": "Short label like 'Fire Safety Certificate', 'Condition Survey', 'As-built drawing', 'BCAR commencement notice', 'Lease agreement', 'Planning permission', 'Room data sheet'. Empty if unclear.",
            },
            "scope_summary": {
                "type": "string",
                "description": "One or two sentence summary of what the document covers.",
            },
            "primary_party": {"type": "string", "description": "The main organisation the document is about or to."},
            "author_firm": {"type": "string", "description": "Firm that prepared/authored the document."},
            "certifier": {"type": "string", "description": "Named certifier, if any."},
            "approved_by": {"type": "string", "description": "Person or body that approved/signed off."},
            "compliance_assertions": {
                "type": "array",
                "items": {"type": "string"},
                "maxItems": 6,
                "description": "Short compliance statements the document makes, verbatim or paraphrased.",
            },
            "issues_identified": {
                "type": "array",
                "items": {"type": "string"},
                "maxItems": 6,
                "description": "Any issues, defects, or non-compliances identified.",
            },
            "date_issued": {"type": "string", "description": "ISO date or year if mentioned."},
            "date_expiry": {"type": "string", "description": "ISO date or year if mentioned."},
            "date_superseded": {"type": "string", "description": "ISO date or year if mentioned."},
            "monetary_amounts": {
                "type": "array",
                "items": {"type": "string"},
                "maxItems": 5,
                "description": "Labelled amounts, e.g. 'Rent: €12,000/year' or 'Total contract sum: €450,000'.",
            },
            "referenced_documents": {
                "type": "array",
                "items": {"type": "string"},
                "maxItems": 5,
                "description": "Other document titles/refs this one cites.",
            },
            "building_address": {"type": "string"},
            "floor_area_m2": {"type": "string"},
            "storeys": {"type": "string"},
            "bed_count": {"type": "string"},
            "rooms_named": {
                "type": "array",
                "items": {"type": "string"},
                "maxItems": 8,
                "description": "Named rooms or spaces, e.g. 'Day room', 'Plant room'.",
            },
            "regulatory_references": {
                "type": "array",
                "items": {"type": "string"},
                "maxItems": 6,
            },
            "site_names": {"type": "array", "items": {"type": "string"}, "maxItems": 4},
            "project_references": {"type": "array", "items": {"type": "string"}, "maxItems": 4},
            "confidence": {
                "type": "string",
                "enum": ["high", "medium", "low"],
                "description": "Self-rated confidence that fields are accurate from the given text.",
            },
            "notes": {"type": "string"},
        },
        "required": [
            "document_type",
            "scope_summary",
            "primary_party",
            "author_firm",
            "certifier",
            "approved_by",
            "compliance_assertions",
            "issues_identified",
            "date_issued",
            "date_expiry",
            "date_superseded",
            "monetary_amounts",
            "referenced_documents",
            "building_address",
            "floor_area_m2",
            "storeys",
            "bed_count",
            "rooms_named",
            "regulatory_references",
            "site_names",
            "project_references",
            "confidence",
            "notes",
        ],
    },
}

COLS = [
    "site_code", "relative_path", "file_name", "extension",
    "document_type", "scope_summary", "primary_party", "author_firm",
    "certifier", "approved_by", "compliance_assertions", "issues_identified",
    "date_issued", "date_expiry", "date_superseded", "monetary_amounts",
    "referenced_documents", "building_address", "floor_area_m2", "storeys",
    "bed_count", "rooms_named", "regulatory_references", "site_names",
    "project_references", "confidence", "notes",
    "model", "input_tokens", "output_tokens", "cost_eur", "status",
]


def extract_text_for(job: dict) -> str:
    """
    Re-run lightweight text extraction on a file. We don't trust whatever was
    left in ContentExtraction.csv because that only stored regex hits, not the
    raw text. Keep output small (<= 12k chars) so input tokens stay cheap.
    """
    path = DNREF_ROOT / job["folder_name"] / job["relative_path"]
    ext = (job.get("extension") or "").lower()
    text = ""
    try:
        if ext == "pdf":
            import pypdf
            r = pypdf.PdfReader(str(path), strict=False)
            for p in r.pages[:3]:
                try:
                    text += (p.extract_text() or "")[:4000] + "\n"
                except Exception:
                    pass
        elif ext == "docx":
            import docx
            d = docx.Document(str(path))
            text = "\n".join(p.text for p in d.paragraphs[:200])
        elif ext == "xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            for sheet in wb.sheetnames[:3]:
                for i, row in enumerate(wb[sheet].iter_rows(values_only=True)):
                    if i > 60:
                        break
                    text += " ".join(str(v) for v in row if v is not None) + "\n"
    except Exception:
        return ""
    return text[:12000]


def already_done_keys():
    done = set()
    if not RESULTS_CSV.exists():
        return done
    with open(RESULTS_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if (row.get("status") or "").lower() == "ok":
                done.add(f"{row['site_code']}::{row['relative_path']}")
    return done


def load_job_index():
    """Map (site_code, relative_path) -> FileInventory row, so we can resolve on-disk paths."""
    idx = {}
    if not INVENTORY_CSV.exists():
        return idx
    with open(INVENTORY_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            key = f"{row['site_code']}::{row['relative_path']}"
            idx[key] = row
    return idx


def load_extract_jobs(done: set):
    """Return ContentExtraction rows with status=ok that aren't already in TierB output."""
    jobs = []
    if not EXTRACT_CSV.exists():
        print(f"ERROR: {EXTRACT_CSV} not found. Run content_extract.py first.", file=sys.stderr)
        sys.exit(1)
    with open(EXTRACT_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if (row.get("status") or "").lower() != "ok":
                continue
            key = f"{row['site_code']}::{row['relative_path']}"
            if key in done:
                continue
            jobs.append(row)
    return jobs


def call_claude(client, model: str, text: str):
    resp = client.messages.create(
        model=model,
        max_tokens=1200,
        system=SYSTEM_PROMPT,
        tools=[TOOL],
        tool_choice={"type": "tool", "name": "record_document"},
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": f"Document text follows.\n\n---\n{text}\n---\n\nExtract the structured record."},
                ],
            }
        ],
    )
    tool_use = None
    for block in resp.content:
        if getattr(block, "type", None) == "tool_use":
            tool_use = block
            break
    if tool_use is None:
        raise RuntimeError("model did not return a tool_use block")
    usage = {
        "input_tokens": getattr(resp.usage, "input_tokens", 0),
        "output_tokens": getattr(resp.usage, "output_tokens", 0),
    }
    return tool_use.input, usage


def main():
    ap = argparse.ArgumentParser(description="Optional paid: Claude structured extraction (Tier B).")
    ap.add_argument(
        "--model",
        default="claude-haiku-4-5-20251001",
        help="Anthropic model id (default: claude-haiku-4-5-20251001).",
    )
    ap.add_argument("--limit", type=int, default=0, help="Process at most N docs (0 = no limit).")
    ap.add_argument("--dry-run", action="store_true", help="Count eligible docs and estimate cost; make no API calls.")
    args = ap.parse_args()

    done = already_done_keys()
    jobs = load_extract_jobs(done)
    if args.limit:
        jobs = jobs[: args.limit]

    est_cost = len(jobs) * COST_PER_DOC_EUR
    print(f"Extracted docs (status=ok): {len(jobs) + len(done):,}")
    print(f"Already in Tier B output:  {len(done):,}")
    print(f"Pending this run:          {len(jobs):,}")
    print(f"Estimated cost:            ~€{est_cost:,.2f}  (@ {COST_PER_DOC_EUR:.3f}/doc on {args.model})")

    if args.dry_run or not jobs:
        return

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY not set.", file=sys.stderr)
        sys.exit(2)

    try:
        import anthropic  # noqa
    except ImportError:
        print("ERROR: anthropic SDK not installed. Run: pip install anthropic", file=sys.stderr)
        sys.exit(2)

    client = anthropic.Anthropic(api_key=api_key)
    inventory = load_job_index()

    mode = "a" if RESULTS_CSV.exists() and RESULTS_CSV.stat().st_size > 0 else "w"
    fh = open(RESULTS_CSV, mode, newline="", encoding="utf-8")
    w = csv.writer(fh)
    if mode == "w":
        w.writerow(COLS)

    cnt = 0
    errors = 0
    skipped = 0
    start = time.time()
    for ce_row in jobs:
        cnt += 1
        key = f"{ce_row['site_code']}::{ce_row['relative_path']}"
        inv = inventory.get(key)
        if not inv:
            skipped += 1
            continue
        text = extract_text_for(inv)
        if not text.strip():
            skipped += 1
            w.writerow([
                ce_row["site_code"], ce_row["relative_path"], ce_row["file_name"], ce_row["extension"],
            ] + [""] * 23 + [args.model, 0, 0, 0.0, "no_text"])
            fh.flush()
            continue

        try:
            record, usage = call_claude(client, args.model, text)
        except Exception as e:
            errors += 1
            w.writerow([
                ce_row["site_code"], ce_row["relative_path"], ce_row["file_name"], ce_row["extension"],
            ] + [""] * 22 + [f"error: {str(e)[:120]}"] + [args.model, 0, 0, 0.0, "error"])
            fh.flush()
            continue

        def j(key: str) -> str:
            v = record.get(key, "")
            if isinstance(v, list):
                return "; ".join(str(x) for x in v)
            return str(v)

        w.writerow([
            ce_row["site_code"], ce_row["relative_path"], ce_row["file_name"], ce_row["extension"],
            j("document_type"), j("scope_summary"), j("primary_party"), j("author_firm"),
            j("certifier"), j("approved_by"),
            j("compliance_assertions"), j("issues_identified"),
            j("date_issued"), j("date_expiry"), j("date_superseded"),
            j("monetary_amounts"), j("referenced_documents"),
            j("building_address"), j("floor_area_m2"), j("storeys"),
            j("bed_count"), j("rooms_named"), j("regulatory_references"),
            j("site_names"), j("project_references"), j("confidence"), j("notes"),
            args.model,
            usage.get("input_tokens", 0), usage.get("output_tokens", 0),
            round(COST_PER_DOC_EUR, 4),
            "ok",
        ])
        fh.flush()

        if cnt % 25 == 0:
            elapsed = time.time() - start
            rate = cnt / elapsed if elapsed > 0 else 0
            eta = (len(jobs) - cnt) / rate if rate > 0 else 0
            print(
                f"  {cnt}/{len(jobs)}  errors: {errors}  skipped: {skipped}  "
                f"rate: {rate:.2f}/s  ETA: {int(eta // 60)}m{int(eta % 60)}s"
            )

    fh.close()
    print(f"\nDone. {cnt} docs processed ({errors} errors, {skipped} skipped).")


if __name__ == "__main__":
    main()
